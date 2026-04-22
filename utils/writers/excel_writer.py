import os
import pandas as pd
from utils.logs.logging import logger


def write_to_excel(data, filename: str, sheet_name: str = 'Data') -> str:
    """
    Exporta datos en formato de libro de cálculo MS Excel preservando
    nativamente la inferencia de tipos de Pandas hacia las celdas de Excel.
    """

    # Conversión al vuelo desde orígenes SQL
    if isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, pd.DataFrame):
        df = data
    else:
        error_msg = f"Estructura incompatible para Excel. Esperado DataFrame o List[dict], recibido: {type(data)}"
        logger.error(error_msg)
        raise ValueError(error_msg)

    if df.empty:
        logger.warning(
            f"La tabla enviada para '{filename}' no posee ningún registro en sus filas.")

    base_dir = os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))))
    output_folder = os.path.join(base_dir, 'data', 'output')
    os.makedirs(output_folder, exist_ok=True)

    # Prevención para que siempre haya extensión
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    file_path = os.path.join(output_folder, filename)

    try:
        # 1. Refinamiento de tipos en Pandas (manejo inteligente de nulos y tipos nativos)
        df = df.convert_dtypes()

        # 2. Volcado a Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # 3. Aplicar formatos de celda específicos y Auto-ajuste
            from openpyxl.utils import get_column_letter
            
            for idx, col in enumerate(df.columns, 1):
                # Determinar el formato de número según el tipo de dato de la columna
                dtype = str(df[col].dtype)
                col_letter = get_column_letter(idx)
                num_format = None
                
                if 'datetime' in dtype:
                    num_format = 'yyyy-mm-dd hh:mm:ss'
                elif 'Int' in dtype or 'int' in dtype:
                    num_format = '0' # Entero puro para evitar separadores de miles en años/IDs
                elif 'Float' in dtype or 'float' in dtype:
                    num_format = None # Excel usará su precisión natural
                
                # Aplicar formato a las celdas de datos
                if num_format:
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=idx).number_format = num_format

                # --- Lógica de Auto-ajuste de ancho ---
                series_lengths = df[col].astype(str).map(len)
                max_val_len = series_lengths.max() if not series_lengths.empty else 0
                max_len = max(max_val_len, len(str(col))) + 3
                adjusted_width = min(max_len, 60) 
                worksheet.column_dimensions[col_letter].width = adjusted_width

        logger.success(f"Excel exportado con preservación de tipos en '{file_path}'")
        return file_path

    except Exception as e:
        logger.exception(
            f"Colapso del motor escribiendo libro de cálculo para '{filename}'")
        raise
